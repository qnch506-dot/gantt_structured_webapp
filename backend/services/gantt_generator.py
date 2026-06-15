from __future__ import annotations

from calendar import monthrange
from io import BytesIO
import re
from datetime import date, timedelta
from typing import Literal

from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
from PIL import Image as PILImage
from PIL import ImageDraw

from .models import ScheduledTask


HEADER_FILL = PatternFill("solid", fgColor="111827")
SUB_HEADER_FILL = PatternFill("solid", fgColor="E5E7EB")
WEEKEND_FILL = PatternFill("solid", fgColor="F3F4F6")
GROUP_FILL = PatternFill("solid", fgColor="F9FAFB")
ERROR_FILL = PatternFill("solid", fgColor="FEE2E2")
MONTH_FILL = PatternFill("solid", fgColor="EEF2FF")

THIN = Side(style="thin", color="D1D5DB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

GROUP_COLOR_PALETTE = [
    "DC2626",  # red
    "2563EB",  # blue
    "16A34A",  # green
    "9333EA",  # purple
    "EA580C",  # orange
    "0891B2",  # cyan
    "BE123C",  # rose
    "4F46E5",  # indigo
    "65A30D",  # lime
    "C026D3",  # fuchsia
]

PeriodMode = Literal["daily", "weekly", "monthly"]


def color_for_group(group_name: str) -> str:
    """
    같은 이름/공종은 항상 같은 색을 갖도록 결정적 색상 매핑을 합니다.
    """
    text = str(group_name or "기본")
    index = sum(ord(ch) for ch in text) % len(GROUP_COLOR_PALETTE)
    return GROUP_COLOR_PALETTE[index]


def hex_to_rgb(color_hex: str) -> tuple[int, int, int]:
    color_hex = color_hex.strip().replace("#", "")
    return int(color_hex[0:2], 16), int(color_hex[2:4], 16), int(color_hex[4:6], 16)


def column_width_to_pixels(width: float | int | None) -> int:
    """
    Excel column width를 대략적인 pixel 값으로 변환합니다.
    openpyxl/Excel의 실제 렌더링과 완전히 같지는 않지만 이미지 폭 계산에 충분합니다.
    """
    if width is None:
        width = 8.43
    return max(int(float(width) * 7 + 5), 8)


def row_height_to_pixels(height_points: float | int | None) -> int:
    if height_points is None:
        height_points = 24
    return max(int(float(height_points) * 96 / 72), 18)


def pixels_to_EMU(px: int) -> int:
    # 1 px ≈ 9525 EMU
    return int(px * 9525)



def make_arrow_png(
    color_hex: str,
    width_px: int,
    height_px: int,
    milestone: bool = False,
    left_tip_x: int | None = None,
    right_tip_x: int | None = None,
) -> BytesIO:
    """
    투명 배경의 연속 화살표 PNG를 생성합니다.
    Excel 셀 위에 올리는 방식이라 셀마다 선이 끊기지 않습니다.

    left_tip_x / right_tip_x를 지정하면 화살표 머리의 '끝점'을
    이미지 내부의 해당 x 좌표에 정확히 맞춥니다.
    """
    width_px = max(width_px, 18)
    height_px = max(height_px, 18)

    image = PILImage.new("RGBA", (width_px, height_px), (255, 255, 255, 0))
    draw = ImageDraw.Draw(image)
    color = (*hex_to_rgb(color_hex), 255)

    y = height_px // 2
    line_width = 4
    arrow_len = 9
    arrow_half = 5

    if milestone:
        size = 6
        cx = width_px // 2 if left_tip_x is None else max(min(left_tip_x, width_px - 2), 2)
        diamond = [
            (cx, y - size),
            (cx + size, y),
            (cx, y + size),
            (cx - size, y),
        ]
        draw.polygon(diamond, fill=color)
    else:
        # 기본값: 기존처럼 양쪽 끝 가까이
        if left_tip_x is None:
            left_tip_x = 2
        if right_tip_x is None:
            right_tip_x = width_px - 2

        left_tip_x = max(min(left_tip_x, width_px - 2), 2)
        right_tip_x = max(min(right_tip_x, width_px - 2), left_tip_x + 12)

        left_base_x = min(left_tip_x + arrow_len, right_tip_x - 4)
        right_base_x = max(right_tip_x - arrow_len, left_tip_x + 4)

        if right_tip_x - left_tip_x < 18:
            # 아주 짧은 작업은 양방향 화살표 느낌으로 처리
            draw.line([(left_tip_x + 2, y), (right_tip_x - 2, y)], fill=color, width=line_width)
            draw.polygon(
                [(left_tip_x, y), (left_tip_x + arrow_len, y - arrow_half), (left_tip_x + arrow_len, y + arrow_half)],
                fill=color,
            )
            draw.polygon(
                [(right_tip_x, y), (right_tip_x - arrow_len, y - arrow_half), (right_tip_x - arrow_len, y + arrow_half)],
                fill=color,
            )
        else:
            draw.line([(left_base_x, y), (right_base_x, y)], fill=color, width=line_width)
            draw.polygon(
                [(left_tip_x, y), (left_base_x, y - arrow_half), (left_base_x, y + arrow_half)],
                fill=color,
            )
            draw.polygon(
                [(right_tip_x, y), (right_base_x, y - arrow_half), (right_base_x, y + arrow_half)],
                fill=color,
            )

    output = BytesIO()
    image.save(output, format="PNG")
    output.seek(0)
    return output





def add_arrow_image(
    ws,
    start_col: int,
    end_col: int,
    row: int,
    color_hex: str,
    milestone: bool = False,
    mode: PeriodMode = "daily",
):
    """
    지정한 셀 구간 위에 투명 PNG 화살표를 올립니다.

    v16 핵심 수정:
    - 좌우 위치는 초기 방식으로 되돌립니다.
      * 일별: 좌우 3px 여백
      * 주간별: 좌우 4px 여백
      * 월별: 좌우 여백 없음
    - 상하 위치만 보정합니다.
      * 이미지 높이를 줄이고
      * rowOff를 더 아래로 내려서
      * Excel에서 화살표가 위로 붙어 보이지 않게 합니다.
    """
    start_col = max(start_col, 1)
    end_col = max(end_col, start_col)

    total_width_px = 0
    for col in range(start_col, end_col + 1):
        col_letter = get_column_letter(col)
        total_width_px += column_width_to_pixels(ws.column_dimensions[col_letter].width)

    total_height_px = row_height_to_pixels(ws.row_dimensions[row].height)

    # 좌우는 '처음 방식'으로 복원
    if mode == "daily":
        left_pad = 3
        right_pad = 3
    elif mode == "weekly":
        left_pad = 4
        right_pad = 4
    else:  # monthly
        left_pad = 0
        right_pad = 0

    # 상하만 조정
    # Excel에서 실제로는 위쪽으로 조금 떠 보이므로 시각적으로 더 아래로 내립니다.
    inner_height_px = min(16, max(total_height_px - 8, 12))
    top_pad = max((total_height_px - inner_height_px) // 2 + 5, 0)

    inner_width_px = max(total_width_px - left_pad - right_pad, 10)

    png = make_arrow_png(
        color_hex=color_hex,
        width_px=inner_width_px,
        height_px=inner_height_px,
        milestone=milestone,
    )

    img = ExcelImage(png)
    img.width = inner_width_px
    img.height = inner_height_px

    marker = AnchorMarker(
        col=start_col - 1,  # 0-based
        colOff=pixels_to_EMU(left_pad),
        row=row - 1,        # 0-based
        rowOff=pixels_to_EMU(top_pad),
    )
    img.anchor = OneCellAnchor(
        _from=marker,
        ext=XDRPositiveSize2D(
            cx=pixels_to_EMU(inner_width_px),
            cy=pixels_to_EMU(inner_height_px),
        ),
    )
    ws.add_image(img)


def daterange(start: date, end: date) -> list[date]:
    days = []
    cur = start
    while cur <= end:
        days.append(cur)
        cur += timedelta(days=1)
    return days


def week_start(d: date) -> date:
    return d - timedelta(days=d.weekday())


def month_start(d: date) -> date:
    return date(d.year, d.month, 1)


def month_end(d: date) -> date:
    last_day = monthrange(d.year, d.month)[1]
    return date(d.year, d.month, last_day)


def build_periods(min_date: date, max_date: date, mode: PeriodMode) -> list[dict]:
    periods: list[dict] = []

    if mode == "daily":
        for d in daterange(min_date, max_date):
            periods.append({
                "start": d,
                "end": d,
                "top": f"{d.year}.{d.month:02d}" if d.day == 1 or d == min_date else "",
                "label": str(d.day),
                "width": 3.8,
            })
        return periods

    if mode == "weekly":
        cur = week_start(min_date)
        idx = 1
        while cur <= max_date:
            end = cur + timedelta(days=6)
            visible_start = max(cur, min_date)
            visible_end = min(end, max_date)
            periods.append({
                "start": cur,
                "end": end,
                "top": f"{visible_start.year}.{visible_start.month:02d}",
                "label": f"{idx}주\n{visible_start.month:02d}/{visible_start.day:02d}~{visible_end.month:02d}/{visible_end.day:02d}",
                "width": 9.8,
            })
            cur += timedelta(days=7)
            idx += 1
        return periods

    cur = month_start(min_date)
    while cur <= max_date:
        end = month_end(cur)
        periods.append({
            "start": cur,
            "end": end,
            "top": str(cur.year),
            "label": f"{cur.month}월",
            "width": 10.8,
        })
        if cur.month == 12:
            cur = date(cur.year + 1, 1, 1)
        else:
            cur = date(cur.year, cur.month + 1, 1)

    return periods


def overlaps(task_start: date, task_end: date, period_start: date, period_end: date) -> bool:
    return task_start <= period_end and task_end >= period_start


def contains(period_start: date, period_end: date, target: date) -> bool:
    return period_start <= target <= period_end


def build_group_ranges(tasks: list[ScheduledTask], start_row: int) -> list[dict]:
    ranges: list[dict] = []
    if not tasks:
        return ranges

    current_group = tasks[0].group
    group_start_index = 0

    for idx, task in enumerate(tasks[1:], start=1):
        if task.group != current_group:
            ranges.append({
                "group": current_group,
                "start_row": start_row + group_start_index,
                "end_row": start_row + idx - 1,
            })
            current_group = task.group
            group_start_index = idx

    ranges.append({
        "group": current_group,
        "start_row": start_row + group_start_index,
        "end_row": start_row + len(tasks) - 1,
    })
    return ranges


def group_task_number(tasks: list[ScheduledTask], current_index: int) -> int:
    group = tasks[current_index].group
    count = 0
    for idx in range(0, current_index + 1):
        if tasks[idx].group == group:
            count += 1
    return count


def format_task_name(tasks: list[ScheduledTask], current_index: int) -> str:
    name = str(tasks[current_index].task_name or "").strip()
    if re.match(r"^\d+\s*[\.\)]\s*", name):
        return name

    number = group_task_number(tasks, current_index)
    return f"{number}. {name}"


def create_timeline_sheet(
    wb: Workbook,
    title: str,
    tasks: list[ScheduledTask],
    mode: PeriodMode,
    active: bool = False,
):
    valid_tasks = [t for t in tasks if t.calculated_start and t.calculated_end]
    if not valid_tasks:
        raise ValueError("간트차트를 만들 수 있는 정상 작업이 없습니다.")

    if active:
        ws = wb.active
        ws.title = title
    else:
        ws = wb.create_sheet(title=title)

    min_date = min(t.calculated_start for t in valid_tasks if t.calculated_start)
    max_date = max(t.calculated_end for t in valid_tasks if t.calculated_end)
    periods = build_periods(min_date, max_date, mode)

    ws.freeze_panes = "F5"
    ws.sheet_view.showGridLines = True

    max_col = max(8, 5 + len(periods))

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = f"공정 간트차트 - {title}"
    title_cell.font = Font(size=18, bold=True, color="111827")
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max_col)
    desc = ws.cell(row=2, column=1)
    desc.value = f"기간: {min_date.isoformat()} ~ {max_date.isoformat()} / 작업 수: {len(valid_tasks)}"
    desc.font = Font(size=10, color="6B7280")
    desc.alignment = Alignment(horizontal="left", vertical="center")

    headers = ["이름", "작업명", "작업일", "시작일", "종료일"]
    for col, label in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col)
        cell.value = label
        cell.fill = HEADER_FILL
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER

    for idx, period in enumerate(periods, start=6):
        top_cell = ws.cell(row=3, column=idx)
        top_cell.value = period["top"]
        top_cell.fill = MONTH_FILL if mode != "daily" else SUB_HEADER_FILL
        top_cell.font = Font(size=9, bold=True, color="374151")
        top_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        top_cell.border = BORDER

        label_cell = ws.cell(row=4, column=idx)
        label_cell.value = period["label"]
        label_cell.fill = HEADER_FILL
        label_cell.font = Font(size=8, color="FFFFFF", bold=True)
        label_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        label_cell.border = BORDER

        ws.column_dimensions[get_column_letter(idx)].width = period["width"]

    start_row = 5

    for task_index, task in enumerate(valid_tasks):
        r = start_row + task_index
        group_arrow_color = color_for_group(task.group)

        # 이미지가 셀 중앙에 들어오도록 행 높이를 먼저 지정합니다.
        ws.row_dimensions[r].height = 28

        left_values = [
            task.group,
            format_task_name(valid_tasks, task_index),
            task.work_days,
            task.calculated_start.isoformat(),
            task.calculated_end.isoformat(),
        ]

        for c, value in enumerate(left_values, start=1):
            cell = ws.cell(row=r, column=c)
            cell.value = value
            cell.border = BORDER

            if c == 1:
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.fill = GROUP_FILL
                cell.font = Font(bold=True, color="111827", size=9)
            elif c == 2:
                cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
                cell.font = Font(color="111827", size=9)
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = Font(color="374151", size=9)

            if task.status == "오류":
                cell.fill = ERROR_FILL

        overlapping_period_indices: list[int] = []

        # 성능 최적화: 대용량 파일에서는 모든 날짜 셀에 border/fill을 칠하지 않습니다.
        # Excel 기본 gridline을 사용하고, 실제 작업 구간은 PNG 화살표만 올립니다.
        for period_index, period in enumerate(periods):
            p_start = period["start"]
            p_end = period["end"]

            if overlaps(task.calculated_start, task.calculated_end, p_start, p_end):
                overlapping_period_indices.append(period_index)

        if overlapping_period_indices:
            if task.work_days == 0:
                milestone_period = None
                for period_index, period in enumerate(periods):
                    if contains(period["start"], period["end"], task.calculated_start):
                        milestone_period = period_index
                        break

                if milestone_period is None:
                    milestone_period = overlapping_period_indices[0]

                col = 6 + milestone_period
                add_arrow_image(
                    ws=ws,
                    start_col=col,
                    end_col=col,
                    row=r,
                    color_hex=group_arrow_color,
                    milestone=True,
                    mode=mode,
                )
            else:
                start_col = 6 + overlapping_period_indices[0]
                end_col = 6 + overlapping_period_indices[-1]
                add_arrow_image(
                    ws=ws,
                    start_col=start_col,
                    end_col=end_col,
                    row=r,
                    color_hex=group_arrow_color,
                    milestone=False,
                    mode=mode,
                )

    # 같은 이름/공종은 A열을 세로 병합해서 공정표 좌측 구조처럼 표시합니다.
    for group_range in build_group_ranges(valid_tasks, start_row):
        if group_range["start_row"] < group_range["end_row"]:
            ws.merge_cells(
                start_row=group_range["start_row"],
                start_column=1,
                end_row=group_range["end_row"],
                end_column=1,
            )

        merged_cell = ws.cell(row=group_range["start_row"], column=1)
        merged_cell.value = group_range["group"]
        merged_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        merged_cell.font = Font(bold=True, color="111827", size=9)
        merged_cell.fill = GROUP_FILL

        for rr in range(group_range["start_row"], group_range["end_row"] + 1):
            ws.cell(row=rr, column=1).border = BORDER

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 36
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14

    # 필터는 좌측 정보 컬럼에만 적용합니다.
    # 날짜/주간/월간 헤더까지 필터를 걸면 각 날짜 칸에 검은 ▼가 생겨서 공정표가 지저분해집니다.
    ws.auto_filter.ref = f"A4:E{start_row + len(valid_tasks) - 1}"

    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A3
    ws.page_margins = PageMargins(left=0.2, right=0.2, top=0.4, bottom=0.4)
    ws.print_title_rows = "1:4"

    return ws


def create_gantt_workbook(tasks: list[ScheduledTask]) -> Workbook:
    """
    Excel 결과 파일에 일별/주간별/월간별 시트를 모두 생성합니다.
    - 일별: 하루 = 한 칸
    - 주간별: 1주 = 한 칸
    - 월간별: 1개월 = 한 칸

    모든 시트의 화살표는 셀마다 문자를 넣는 방식이 아니라,
    투명 PNG를 셀 위에 올려 하나의 연속된 선처럼 보이게 합니다.
    """
    wb = Workbook()
    create_timeline_sheet(wb, "일별", tasks, "daily", active=True)
    create_timeline_sheet(wb, "주간별", tasks, "weekly")
    create_timeline_sheet(wb, "월간별", tasks, "monthly")
    return wb
