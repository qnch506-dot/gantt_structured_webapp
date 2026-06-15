from __future__ import annotations

from datetime import date, datetime, timedelta
from io import BytesIO
from typing import Any, Dict, Optional

from openpyxl import load_workbook

from .models import RawTask


HEADER_ALIASES = {
    "group": ["이름", "분류", "공종", "그룹", "category", "group", "name"],
    "task_name": ["작업명", "세부작업", "공정명", "task", "task_name", "work"],
    "work_days": ["작업일", "기간", "일수", "duration", "work_days", "days"],
    "start_date": ["시작일", "시작일자", "착수일", "start", "start_date"],
    "end_date": ["종료일", "종료일자", "완료일", "end", "end_date", "finish"],
    "progress": ["진행률", "공정률", "progress", "rate"],
    "predecessor": ["선행작업", "선행", "predecessor", "dependency"],
    "floor": ["층", "floor"],
    "zone": ["구역", "zone", "area"],
}


def normalize_header(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip().lower().replace(" ", "").replace("_", "")


SKIP_SHEET_KEYWORDS = [
    "설명",
    "요약",
    "작업수",
    "피벗",
    "통계",
    "summary",
    "readme",
    "pivot",
]


def should_skip_sheet(sheet_name: str) -> bool:
    normalized = str(sheet_name or "").strip().lower()
    return any(keyword.lower() in normalized for keyword in SKIP_SHEET_KEYWORDS)


def find_header_row(ws, max_scan_rows: int = 10) -> tuple[int, Dict[str, int]]:
    """
    상단 10행 안에서 실제 작업 데이터 헤더를 찾습니다.

    중요한 점:
    - 작업 데이터가 아닌 설명/요약/공종별 집계 시트는 읽지 않도록 빈 mapping을 반환합니다.
    - `작업명`이 없거나, `작업일/시작일/종료일` 중 하나도 없으면 작업 시트로 보지 않습니다.
    - 예전 `작업명 / 기간` 2열 구조는 계속 지원합니다.
    """
    if should_skip_sheet(ws.title):
        return 1, {}

    alias_map: Dict[str, str] = {}
    for key, aliases in HEADER_ALIASES.items():
        for alias in aliases:
            alias_map[normalize_header(alias)] = key

    best_row = 1
    best_mapping: Dict[str, int] = {}
    best_score = 0

    for row in range(1, min(max_scan_rows, ws.max_row) + 1):
        mapping: Dict[str, int] = {}
        for col in range(1, ws.max_column + 1):
            cell_value = normalize_header(ws.cell(row=row, column=col).value)
            if cell_value in alias_map:
                mapping[alias_map[cell_value]] = col

        # 작업명은 필수로 보고, 날짜 또는 작업일 관련 컬럼이 있어야 실제 작업 데이터로 인정합니다.
        has_task_name = "task_name" in mapping
        has_schedule_basis = any(key in mapping for key in ["work_days", "start_date", "end_date"])
        score = len(mapping) if has_task_name and has_schedule_basis else 0

        if score > best_score:
            best_score = score
            best_row = row
            best_mapping = mapping

    if best_score == 0:
        return 1, {}

    # A열 작업명/B열 기간만 있는 옛 백데이터도 지원
    if ws.max_column >= 2:
        a = normalize_header(ws.cell(row=best_row, column=1).value)
        b = normalize_header(ws.cell(row=best_row, column=2).value)
        if a in [normalize_header(x) for x in HEADER_ALIASES["task_name"]] and b in [normalize_header(x) for x in HEADER_ALIASES["work_days"]]:
            best_mapping = {
                "task_name": 1,
                "work_days": 2,
            }

    return best_row, best_mapping


def parse_int(value: Any) -> Optional[int]:
    if value is None or value == "":
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(round(value))
    text = str(value).strip().replace("일", "").replace(",", "")
    if not text:
        return None
    try:
        return int(float(text))
    except ValueError:
        return None


def parse_date(value: Any) -> Optional[date]:
    if value is None or value == "":
        return None

    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    # Excel serial date 보정
    if isinstance(value, (int, float)):
        # openpyxl이 보통 날짜 타입으로 반환하지만, 숫자로 온 경우 대비
        try:
            return date(1899, 12, 30) + timedelta(days=int(value))
        except Exception:
            return None

    text = str(value).strip()
    if not text:
        return None

    # 흔한 한국식 날짜 포맷 대응
    candidates = [
        "%Y-%m-%d",
        "%Y.%m.%d",
        "%Y/%m/%d",
        "%y-%m-%d",
        "%y.%m.%d",
        "%y/%m/%d",
    ]
    for fmt in candidates:
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass

    return None


def parse_progress(value: Any) -> Optional[float]:
    if value is None or value == "":
        return None

    if isinstance(value, (int, float)):
        number = float(value)
        return number * 100 if 0 <= number <= 1 else number

    text = str(value).strip().replace("%", "")
    if not text:
        return None

    try:
        number = float(text)
        return number * 100 if 0 <= number <= 1 else number
    except ValueError:
        return None


META_ROW_KEYWORDS = [
    "포함 컬럼",
    "제거 컬럼",
    "권장 시트",
    "웹 업로드",
    "주의",
    "안내",
    "설명",
    "비고",
    "메모",
    "참고",
    "사용법",
    "예시",
    "업로드용",
    "통합 업로드용",
]

HEADER_WORDS_IN_TEXT = [
    "이름",
    "작업명",
    "작업일",
    "시작일",
    "종료일",
    "금액",
    "보험",
    "진행률",
    "비고",
]


def looks_like_instruction_row(group_value: Any, task_name_value: Any, work_days_value: Any, start_value: Any, end_value: Any) -> bool:
    """
    실제 공정 작업이 아니라, 엑셀 안에 적어둔 안내/설명/주의 문구 행을 제외합니다.

    예:
    - 포함 컬럼 / 이름, 작업명, 작업일, 시작일, 종료일
    - 제거 컬럼 / 금액, 보험, 진행률, 비고
    - 웹 업로드 권장 시트 / 통합 업로드용
    - 주의 / 이미지 판독 기반이므로...
    """
    group_text = str(group_value or "").strip()
    task_text = str(task_name_value or "").strip()
    combined = f"{group_text} {task_text}"

    # 날짜/작업일이 하나라도 있으면 실제 작업 후보로 봅니다.
    if work_days_value or start_value or end_value:
        return False

    if not combined:
        return False

    for keyword in META_ROW_KEYWORDS:
        if keyword in combined:
            return True

    # "이름, 작업명, 작업일, 시작일, 종료일"처럼 헤더 목록을 설명하는 행 제외
    header_word_count = sum(1 for word in HEADER_WORDS_IN_TEXT if word in task_text)
    if header_word_count >= 3:
        return True

    # 쉼표로 여러 컬럼명을 나열한 행 제외
    if "," in task_text or "，" in task_text:
        header_like_count = sum(1 for word in HEADER_WORDS_IN_TEXT if word in task_text)
        if header_like_count >= 2:
            return True

    return False


def cell(ws, row: int, mapping: Dict[str, int], key: str) -> Any:
    col = mapping.get(key)
    if not col:
        return None
    return ws.cell(row=row, column=col).value


def read_tasks_from_excel(file_obj: BytesIO) -> list[RawTask]:
    wb = load_workbook(file_obj, data_only=True)
    tasks: list[RawTask] = []

    for ws in wb.worksheets:
        header_row, mapping = find_header_row(ws)

        # 작업 데이터 시트가 아니면 건너뜁니다.
        if not mapping:
            continue

        for row in range(header_row + 1, ws.max_row + 1):
            task_name_raw = cell(ws, row, mapping, "task_name")
            work_days_raw = cell(ws, row, mapping, "work_days")
            start_raw = cell(ws, row, mapping, "start_date")
            end_raw = cell(ws, row, mapping, "end_date")

            # 완전 빈 줄 스킵
            if not any([task_name_raw, work_days_raw, start_raw, end_raw]):
                continue

            group_raw = cell(ws, row, mapping, "group")

            if looks_like_instruction_row(group_raw, task_name_raw, work_days_raw, start_raw, end_raw):
                continue

            task_name = str(task_name_raw).strip() if task_name_raw else ""
            if not task_name:
                continue

            group = str(group_raw).strip() if group_raw else ws.title

            tasks.append(
                RawTask(
                    row_no=row,
                    group=group,
                    task_name=task_name,
                    work_days=parse_int(work_days_raw),
                    start_date=parse_date(start_raw),
                    end_date=parse_date(end_raw),
                    progress=parse_progress(cell(ws, row, mapping, "progress")),
                    predecessor=str(cell(ws, row, mapping, "predecessor")).strip() if cell(ws, row, mapping, "predecessor") else None,
                    floor=str(cell(ws, row, mapping, "floor")).strip() if cell(ws, row, mapping, "floor") else None,
                    zone=str(cell(ws, row, mapping, "zone")).strip() if cell(ws, row, mapping, "zone") else None,
                )
            )

    if not tasks:
        raise ValueError("읽을 수 있는 작업 데이터가 없습니다. 최소 작업명/작업일 또는 작업명/시작일/종료일이 필요합니다.")

    return tasks
